from openpyxlplus import cell_range
from openpyxl.utils.dataframe import dataframe_to_rows
from numpy import ndarray,array,newaxis
from openpyxl.styles import Alignment

def write_value(data,ws,cell=None,keep_style=True):
    """
    Write value to a cell.

    Parameters:
    data: data to write to cell.
    ws: openpyxl.worksheet.worksheet.Worksheet object
    cell: openpyxl.cell.cell.Cell object. Default to A1 cell of ws.
    keep_style: Whether to preserve original cell style

    Return:
    Returns SheetCellRange object, which allows modification to where the data 
        is written to.
    """
    if cell is None:
        cell = ws["A1"]
    rg = cell_range.SheetCellRange(ws,range_string=cell.coordinate)\
        .write(data,keep_style=keep_style)
    return(rg)
    

def write_list(data,ws,cell=None,direction="down",keep_style=True):
    """
    Write data to row or column

    Parameters:
    data: list,tuple or other iterable objects. Does NOT support nested list.
    ws: openpyxl.worksheet.worksheet.Worksheet object
    cell: openpyxl.cell.cell.Cell object. Cell to start writing. Default to 
        A1 cell of ws.
    direction: one of "up","down","left" or "right". From the starting cell, 
        direction to write data to.
    keep_style: Whether to preserve original cell styles.

    Return:
    Returns SheetCellRange object, which allows modification to where the data 
        is written to.
    """
    if cell is None:
        cell = ws["A1"]
    length = len(data)
    row, col = cell.row, cell.column
    if direction == "down":
        min_col = col
        min_row = row
        max_col = col
        max_row = row+length-1
        data = array(data)[newaxis].transpose()
    elif direction == "up":
        min_col = col
        min_row = row-length+1
        if min_row < 1:
            raise ValueError("Min row is fewer than 1. Increase starting row.")
        max_col = col
        max_row = row
        data = data[::-1]
        data = array(data)[newaxis].transpose()
    elif direction == "right":
        min_col = col
        min_row = row
        max_col = col + length - 1
        max_row = row
    elif direction == "left":
        min_col = col - length + 1
        if min_col < 1:
            raise ValueError(
                "Min column is fewer than 1. Increase starting column."
            )
        min_row = row
        max_col = col
        max_row = row
        data = data[::-1]
    else:
        raise ValueError(f"{direction} is not one of up,down,left or right.")
    rg = cell_range.SheetCellRange(ws,min_col=min_col,min_row=min_row,
        max_col=max_col,max_row=max_row).write(data,keep_style=keep_style)
    return(rg)

def write_array(data,ws,cell=None,keep_style=True):
    """
    Write data to range starting at provided cell.

    Parameters:
    data: List of list or numpy array. If data is list of list, length of all 
        nested lists must be equal.
    ws: openpyxl.worksheet.worksheet.Worksheet object
    cell: openpyxl.cell.cell.Cell object. Cell to start writing, this cell is 
        the top left cell in array. Default to first A1 of ws.
    keep_style: Whether to preserve original cell styles.

    Return:
    Returns SheetCellRange object, which allows modification to where the data 
        is written to.
    """
    if cell is None:
        cell = ws["A1"]
    width = len(data[0])
    height = len(data)
    if type(data) != ndarray:
        for i,x in enumerate(data):
            if len(x) != width:
                raise ValueError(
                    f"Length of nested {str(type(x))} mismatch: "
                    f"length of {x} ({len(x)}) is not equal to {width}."
                )

    min_col = cell.column
    min_row = cell.row
    max_col = cell.column + width - 1
    max_row = cell.row + height - 1

    rg = cell_range.SheetCellRange(ws,min_col=min_col,min_row=min_row,
        max_col=max_col,max_row=max_row).write(data,keep_style=keep_style)
    return(rg)


def write_value_merged(
    data,
    ws,
    cell=None,
    right = 0,
    down = 0,
    left = 0,
    up = 0,
    center=True,
    keep_style=True
):
    """
    Write value to given cell and merge cells with shape defined by shape
    parameter

    Parameters:
    data: scalar value to write to worksheet
    ws: openpyxl.worksheet.worksheet.Worksheet object
    cell: openpyxl.cell.cell.Cell object. Cell to start writing, this cell is 
        the top left cell in array. Default to first A1 of ws.
    right,down,left,up: How many cells to expand in given direction.
    center: Whether to center value in merged cell range
    keep_style: Whether to preserve original cell styles.
    """
    if cell is None:
        cell = ws["A1"]
    rg = cell_range.SheetCellRange(ws,range_string=cell.coordinate)\
        .write(data,keep_style=keep_style)

    rg.expand(right=right,down=down,left=left,up=up)
    rg.cells.set_value(data)
    if center:
        rg.cells.set_style(
            "alignment",
            Alignment(horizontal='center',vertical="center")
        )
    rg.merge_cells()
    return(rg)

def write_dataframe(data,ws,cell=None,index=False,header=True,keep_style=True):
    """
    Write pandas data frame to range starting at provided cell.

    Note: Bear in mind that openpyxl.utils.dataframe.dataframe_to_rows adds 
    index name when index = True. This may cause confusion sometimes.

    Parameters:
    data: pandas dataframe.
    ws: openpyxl.worksheet.worksheet.Worksheet object
    cell: openpyxl.cell.cell.Cell object. Cell to start writing, this cell is 
        the top left cell in array. Default to first A1 of ws.
    index: Whether to write index
    header: Whether to write header
    keep_style: Whether to preserve original cell styles.

    Return:
    Returns TableRange object, which allows modification to where the data is 
        written to.
    """
    if cell is None:
        cell = ws["A1"]
    index_nlevels = data.index.nlevels if index else 0
    header_nlevels = data.columns.nlevels if header else 0

    row_anchor = cell.row
    col_anchor = cell.column

    for i,row in enumerate(dataframe_to_rows(data,index=index,header=header)):
        for j, cell in enumerate(row):
            row_num = row_anchor + i
            col_num = col_anchor + j
            if keep_style:
                from copy import copy
                original_style = copy(ws.cell(row_num,col_num)._style)
                ws.cell(row_num,col_num,value=cell)
                ws.cell(row_num,col_num)._style = original_style
            else:
                ws.cell(row_num,col_num,value=cell)
    # table range:
    min_col = col_anchor
    min_row = row_anchor
    max_col = min_col + j
    max_row = min_row + i
    n_header = header_nlevels + int(index) if index else header_nlevels
    table_range = cell_range.TableRange(
        ws,
        min_col=min_col,
        min_row=min_row,
        max_col=max_col, 
        max_row=max_row,
        n_index=index_nlevels,
        n_header=n_header
    )
    return(table_range)