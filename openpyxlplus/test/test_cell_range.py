import unittest
from openpyxlplus.cell_range import SheetCellRange,SheetTableRange,Cells
from openpyxlplus import writer
from openpyxl import load_workbook,Workbook
from openpyxl.styles import Font,Border,Side
import numpy as np
import os
current_folder = os.path.dirname(__file__)
wb = load_workbook(current_folder +"/test_cell_range.xlsx")

class TestSheetCellRange(unittest.TestCase):
    def setUp(self):
        self.wb = load_workbook(current_folder +"/test_cell_range.xlsx")
        self.ws = wb["Sheet1"]
        self.sheet_range = SheetCellRange(self.ws,range_string="B2:G5")
    def tearDown(self):
        self.wb.close()

    def test_construct(self):
        # construct using min/max col/row
        range1 = SheetCellRange(self.ws,min_col=2,min_row=2,max_col=7,max_row=5)
        self.assertTrue(np.array_equal(
            range1.cells.get_value(),
            self.sheet_range.cells.get_value()
        ))

    def test_cells(self):
        # test if .cells is dynamic
        self.assertEqual(self.sheet_range.cells[0,0].coordinate,"B2")
        self.sheet_range.shift(col_shift = 1)
        self.assertEqual(self.sheet_range.cells[0,0].coordinate,"C2")
        self.sheet_range.shift(col_shift = -1)
        self.assertEqual(self.sheet_range.cells[0,0].coordinate,"B2")

    def test_get_cell(self):
        # get cell using abolute coordinate
        self.assertEqual(self.sheet_range.get_cell((2,3),False).value,"a")
        # get cell using relative coordinate
        self.assertEqual(self.sheet_range.get_cell((0,1)).value,"a")
    
    def test_get_cells(self):
        # doesn't work on single coordinate
        with self.assertRaises(TypeError):
            self.sheet_range.get_cells((2,3))

        # shape is preserved
        # 2d
        coordinates = [
            [(1,1),(1,2)],
            [(2,1),(2,2)]
        ]
        self.assertTrue(np.array_equal(
            self.sheet_range.get_cells(coordinates).get_value(),
            np.array([[1,2],[6,7]])
        ))

        # 1d
        coordinates = [(1,1),(1,2)]
        self.assertTrue(np.array_equal(
            self.sheet_range.get_cells(coordinates).get_value(),
            np.array([1,2])
        ))

    def test_get_subset(self):
        wb = Workbook()
        ws = wb.active
        rg = SheetCellRange(ws,"A1:E5")
        
        # use slice
        self.assertEqual(rg.get_subset("2:,:").coord,"A3:E5") # positive index
        self.assertEqual(rg.get_subset(":,3:").coord,"D1:E5") # positive index
        self.assertEqual(rg.get_subset(":-1,:-1").coord,"A1:D4") # negative index

        # use index
        # works with full index provided
        self.assertEqual(
            rg.get_subset(min_col_idx=1,max_col_idx=4,min_row_idx=1,max_row_idx=4)\
                .coord,
            "B2:D4"
        )

        # works with partial index provided
        self.assertEqual(
            rg.get_subset(min_col_idx=1,max_row_idx=-1).coord,
            "B1:E4"
        )

    def test_clear(self):
        ws = wb["Clear"]
        # only clear value, preserve style
        rg = SheetCellRange(ws,"A1").clear(formatting=False)
        current_cell = rg.get_cell((0,0))
        self.assertIsNone(current_cell.value)
        self.assertEqual(current_cell.font.b,True)
        self.assertEqual(current_cell.fill.start_color.rgb,"FFFFFF00") # backgournd yellow
        self.assertEqual(current_cell.font.color.rgb,"FFFF0000") # font red

        # only clear formatting, keep value
        rg = SheetCellRange(ws,"B1").clear(value=False)
        current_cell = rg.get_cell((0,0))
        self.assertEqual(current_cell.value,"color,bold,fill")
        self.assertEqual(ws["B1"].font.b,False)
        self.assertNotEqual(ws["B1"].fill.start_color.rgb,"FFFFFF00") 
        self.assertTrue("rgb" not in ws["B1"].font.color.__dict__)

    def test_merge_consecutive_cells(self):
            wb = Workbook()
            ws = wb.active
            array = [
                [1,1,2,3],
                [1,2,2,3],
                [3,2,3,3]
            ]
            # merged cells other than the first one (top left) has value of None.
            # the styles of merged cells are stored in the first cell (top left)

            # default on row, center=True
            rg = SheetCellRange(ws,range_string="A1:D3").write(array)
            rg.merge_consecutive_cells()

            self.assertTrue(np.array_equal(
                rg.cell_values,
                [
                    [1,None,2,3],
                    [1,2,None,3],
                    [3,2,3,None]
                ]
            ))

            self.assertTrue(np.array_equal(
                rg.cells.get_style_detail("alignment",["horizontal"]),
                [
                    ['center',None,None,None],
                    [None,'center',None,None],
                    [None,None,'center',None]
                ]
            ))

            self.assertTrue(np.array_equal(
                rg.cells.get_style_detail("alignment","vertical"),
                [
                    ['center',None,None,None],
                    [None,'center',None,None],
                    [None,None,'center',None]
                ]
            ))

            # on column, center=False
            rg = SheetCellRange(ws,range_string="A7:D9").write(array)
            rg.merge_consecutive_cells(on="column",center=False)


            self.assertTrue(np.array_equal(
                rg.cell_values,
                [
                    [1,1,2,3],
                    [None,2,None,None],
                    [3,None,3,None]
                ]
            ))

            self.assertTrue(all(
                rg.cells.get_style_detail("alignment","vertical").flatten() == None
            ))

            self.assertTrue(all(
                rg.cells.get_style_detail("alignment","horizontal").flatten() == None
            ))

            wb.close()

    def test_autosize(self):
        wb = Workbook()
        ws = wb.active
        array = [
            [    "123\n123"     ,  "1\n1"  ,  ""  , "123456"],
            [        "1"        , "12\n12" ,  ""  ,    ""   ],
            [        ""         ,    ""    ,  ""  ,    ""   ],
            ["1\n1\n1\n1\n1\n1" ,     ""   ,  ""  ,    ""   ]
        ]
        fontsize = 1
        rg = writer.write_array(array,ws)
        rg.cells.set_style("font",Font(size=fontsize))
        rg.autosize(
            wrap_text=True,
            min_width = 1 * fontsize,
            min_height = 1 * fontsize,
            max_width= 5 * fontsize,
            max_height = 5 * fontsize,
            width_factor = 1,
            height_factor = 1
        )
        # column A,B,C,D
        self.assertListEqual(
            [int(ws.column_dimensions[x].width) for x in ["A","B","C","D"]],
            [3,2,1,5]
        )

        # row 1,2,3,4
        self.assertListEqual(
            [int(ws.row_dimensions[x].height) for x in [1,2,3,4]],
            [2,2,1,5]
        )
        wb.close()



class TestSheetTableRange(unittest.TestCase):
    def setUp(self):
        self.wb = load_workbook(current_folder +"/test_cell_range.xlsx")
        self.ws = wb["Sheet1"]
    def tearDown(self):
        self.wb.close()
    
    def test_index_header(self):
        ## Single level index and header
        # with index and header
        table_range1 = SheetTableRange(self.ws,range_string="B2:G5",n_index=1,
            n_header=1)
        self.assertEqual(table_range1.header.coord,"C2:G2")
        self.assertEqual(table_range1.index.coord,"B3:B5")
        self.assertEqual(table_range1.body.coord,"C3:G5")
        self.assertEqual(table_range1.top_left_corner.coord,"B2")

        # with header
        table_range1 = SheetTableRange(self.ws,range_string="B2:G5",n_index=0,
            n_header=1)
        self.assertEqual(table_range1.header.coord,"B2:G2")
        self.assertIsNone(table_range1.index)
        self.assertEqual(table_range1.body.coord,"B3:G5")
        self.assertIsNone(table_range1.top_left_corner)

        # with index
        table_range1 = SheetTableRange(self.ws,range_string="B2:G5",n_index=1,
            n_header=0)
        self.assertIsNone(table_range1.header)
        self.assertEqual(table_range1.index.coord,"B2:B5")
        self.assertEqual(table_range1.body.coord,"C2:G5")
        self.assertIsNone(table_range1.top_left_corner)

        # with no index and no header
        table_range1 = SheetTableRange(self.ws,range_string="B2:G5",n_index=0,
            n_header=0)
        self.assertIsNone(table_range1.header)
        self.assertIsNone(table_range1.index)
        self.assertEqual(table_range1.body.coord,"B2:G5")
        self.assertIsNone(table_range1.top_left_corner)

        ## Double level index and header
        # with index and header
        table_range1 = SheetTableRange(self.ws,range_string="A1:G5",n_index=2,
            n_header=2)
        self.assertEqual(table_range1.header.coord,"C1:G2")
        self.assertEqual(table_range1.index.coord,"A3:B5")
        self.assertEqual(table_range1.body.coord,"C3:G5")
        self.assertEqual(table_range1.top_left_corner.coord,"A1:B2")

        # with header
        table_range1 = SheetTableRange(self.ws,range_string="C1:G5",n_index=0,
            n_header=2)
        self.assertEqual(table_range1.header.coord,"C1:G2")
        self.assertIsNone(table_range1.index)
        self.assertEqual(table_range1.body.coord,"C3:G5")
        self.assertIsNone(table_range1.top_left_corner)

        # with index
        table_range1 = SheetTableRange(self.ws,range_string="A3:G5",n_index=2,
            n_header=0)
        self.assertIsNone(table_range1.header)
        self.assertEqual(table_range1.index.coord,"A3:B5")
        self.assertEqual(table_range1.body.coord,"C3:G5")
        self.assertIsNone(table_range1.top_left_corner)

        # with no index and no header
        table_range1 = SheetTableRange(self.ws,range_string="C3:G5",n_index=0,
            n_header=0)
        self.assertIsNone(table_range1.header)
        self.assertIsNone(table_range1.index)
        self.assertEqual(table_range1.body.coord,"C3:G5")
        self.assertIsNone(table_range1.top_left_corner)

class TestCells(unittest.TestCase):
    def setUp(self):
        self.wb = load_workbook(current_folder +"/test_cell_range.xlsx")
        self.ws = wb["Sheet1"]
        # follwoing line tests construction from range
        self.cells = Cells.from_range(self.ws,"B2:G5")
    def tearDown(self):
        self.wb.close()

    def test_get_style(self):
        # shape is preserved
        self.assertTrue(np.array_equal(
            self.cells[:2,:].get_style("font").shape,
            (2,6)
        ))

        # get attribute value correctly
        f = np.vectorize(lambda x: x.b)
        self.assertTrue(np.array_equal(
            f(self.cells[:2,:2].get_style("font")),
            np.array([[False,False],[False,False]])
        ))

    def test_get_style_detail(self):
        # detail in single level
        wb = Workbook()
        ws = wb.active
        cells = Cells.from_range(ws,"A1:B5")

        cells.modify_style("font",Font(b=True))
        cells.add_border()

        self.assertTrue(np.array_equal(
            cells[:,[0]].get_style_detail("font","b"),
            np.array([True]*5).reshape((5,1))
        ))

        # detail in multiple level
        self.assertTrue(np.array_equal(
            cells[:,[0]].get_style_detail("border",["left","style"]),
            np.array(["thin"]*5).reshape((5,1))
        ))

        self.assertTrue(np.array_equal(
            cells[:,[0]].get_style_detail("border","left.style"),
            np.array(["thin"]*5).reshape((5,1))
        ))

    def test_set_style(self):
        # set style with list
        self.cells[:,0].set_style("font",[Font(bold=True)] * 4)
        self.assertTrue(np.array_equal(
            self.cells[:,0].get_style_detail("font","b"),
            [True] * 4
        ))
        # set style with ndarray of same size
        self.cells[:,1:3]\
            .set_style("font",np.array([Font(bold=True)]*8).reshape((4,2)))
        self.assertTrue(np.array_equal(
            self.cells[:,1:3].get_style_detail("font","b"),
            np.array([True] * 8).reshape((4,2))
        ))
        # set style with a single value
        self.cells[:,3]\
            .set_style("font",Font(bold=True))
        self.assertTrue(np.array_equal(
            self.cells[:,3].get_style_detail("font","b"),
            [True] * 4
        ))
        # set style with different shape: fail
        with self.assertRaises(ValueError):
            self.cells[:,3].set_style("font",[Font(bold=True)] * 2)
    
    def test_modify_style(self):
        self.cells[:,0]\
            .modify_style("border",Border(left=Side(style="thin")))\
            .modify_style("border",Border(right=Side(style="thin"))) 
        # left border is thin
        self.assertTrue(np.array_equal(
            self.cells[:,0].get_style_detail("border",["left","style"]),
            ["thin"] * 4
        ))

        # right border is thin
        self.assertTrue(np.array_equal(
            self.cells[:,0].get_style_detail("border",["right","style"]),
            ["thin"] * 4
        ))

    def test_get_value(self):
        self.assertTrue(np.array_equal(
            self.cells[:,1].get_value(),
            np.array(["a",1,6,11],dtype="O")
        ))
        # get value without shape changed
        self.assertTrue(np.array_equal(
            self.cells[:,[1]].get_value(),
            np.array(["a",1,6,11],dtype="O").reshape((4,1))
        ))

    def test_set_value(self):
        wb = load_workbook(current_folder +"/test_cell_range.xlsx")
        ws = wb["Sheet1"]
        cells = Cells.from_range(ws,"B2:G5")

        # set value with list
        cells_1 = cells[1:,1]
        cells_1.set_value([1,1,1])
        self.assertTrue(np.array_equal(
            cells[1:,1].get_value(),
            [1,1,1]
        ))
        # set value with array with same size
        cells_2 = cells[1:,[1]]
        cells_2.set_value(np.array([2,2,2]).reshape((3,1)))
        self.assertTrue(np.array_equal(
            cells[1:,[1]].get_value(),
            np.array([2,2,2]).reshape((3,1))
        ))
        # set value with scaler
        cells_2 = cells[1:,1]
        cells_2.set_value(3)
        self.assertTrue(np.array_equal(
            cells[1:,1].get_value(),
            [3,3,3]
        ))
        # set value with array of different size: fail
        with self.assertRaises(ValueError):
            cells[1:,1].set_value([1,2])

    def test_set_value_preserve_style(self):
        ws = wb["DefaultStyle"]
        # default to preserve style
        SheetCellRange(ws,"A1").cells.set_value(1)
        self.assertEqual(ws["A1"].font.b,True)
        self.assertEqual(ws["A1"].fill.start_color.rgb,"FFFFFF00") # backgournd yellow
        self.assertEqual(ws["A1"].font.color.rgb,"FFFF0000") # font red
        # disable preserve style
        SheetCellRange(ws,"B1").cells.set_value(1,keep_style=False)
        self.assertEqual(ws["B1"].font.b,False)
        self.assertNotEqual(ws["B1"].fill.start_color.rgb,"FFFFFF00") 
        self.assertTrue("rgb" not in ws["B1"].font.color.__dict__) # font color not set (default)

    def test_side(self):
        # head
        self.assertTrue(np.array_equal(
            self.cells.head(1).get_value(),
            np.array([[None,"a","b","c","d","e"]])
        ))

        # tail
        self.assertTrue(np.array_equal(
            self.cells.tail(1).get_value(),
            np.array([["r3",11,12,13,14,15]],dtype="O")
        ))

        # left
        self.assertTrue(np.array_equal(
            self.cells.left(2).get_value(),
            np.array([[None,"a"],["r1",1],["r2",6],["r3",11]],dtype="O")
        ))

        # right
        self.assertTrue(np.array_equal(
            self.cells.right(2).get_value(),
            np.array([["d","e"],[4,5],[9,10],[14,15]],dtype="O")
        ))

    def test_add_border(self):
        wb = load_workbook(current_folder +"/test_cell_range.xlsx")
        ws = wb["Sheet1"]
        cells = Cells.from_range(ws,"B2:G5")
        # range more than 1x1 ----
        cells_subset1 = cells[:2,:2]
        cells_subset1.add_border()
        # left
        self.assertTrue(np.array_equal(
            cells_subset1.left().get_style_detail("border",["left","style"]),
            np.array(["thin"]*2).reshape((2,1))
        ))
        # right
        self.assertTrue(np.array_equal(
            cells_subset1.right().get_style_detail("border",["right","style"]),
            np.array(["thin"]*2).reshape((2,1))
        ))       
        # top
        self.assertTrue(np.array_equal(
            cells_subset1.head().get_style_detail("border",["top","style"]),
            np.array(["thin"]*2).reshape((1,2))
        ))      
        # bottom
        self.assertTrue(np.array_equal(
            cells_subset1.tail().get_style_detail("border",["bottom","style"]),
            np.array(["thin"]*2).reshape((1,2))
        ))

    def test_add_border_w_single_side(self):
        wb = load_workbook(current_folder +"/test_cell_range.xlsx")
        ws = wb["Sheet1"]
        rg = SheetCellRange(ws,"A1:C3")

        four_sides = ["left","right","top","bottom"]

        for side in four_sides:
            key_arguments = {
                "left": side == "left",
                "right": side == "right",
                "top": side == "top",
                "bottom": side == "bottom"
            }
            # also test SheetCellRange.add_border() here
            rg.add_border(**key_arguments)
            
            # selected side is modified
            self.assertTrue(np.array_equal(
                getattr(rg.cells,side)()\
                    .get_style_detail("border",[side,"style"]).flatten(),
                np.array(["thin"]*3).flatten())
            )
            # other sides are not modified
            for side_other in list(set(four_sides) - set([side])):
                self.assertTrue(np.array_equal(
                    getattr(rg.cells,side_other)()\
                        .get_style_detail("border",[side_other,"style"]).flatten(),
                    np.array([None]*3).flatten())
            )
            # reset side to None
            empty_side = Side(style=None)
            rg.cells.set_style(
                "border",
                Border(
                    left=empty_side,right=empty_side,
                    top=empty_side,bottom=empty_side
                )
            )

    def test_mixed_type_cells(self):
        ws = self.wb["Sheet2"]
        # get array with mixed types
        self.assertTrue(np.array_equal(
            SheetCellRange(ws,"B2:D5").cells.get_value(),
            np.array([
                ["int","string","num"],
                [1,"a",0.1],
                [2,"b",0.2],
                [3,"c",0.3]
            ],dtype="O")
        ))
        # can overwrite cell with different type
        SheetCellRange(ws,"B2:D5").cells[:,1].set_value(["test",1,2,3])

        self.assertTrue(np.array_equal(
            SheetCellRange(ws,"B2:D5").cells[:,1].get_value(),
            np.array(["test",1,2,3],dtype="O")
        ))
        # can write with 2D array
        self.assertTrue(np.array_equal(
            SheetCellRange(ws,"C2:C5").cells.get_value(),
            np.array([["test"],[1],[2],[3]],dtype="O")
        )) 
    def test_to_range(self):
        wb = Workbook()
        ws = wb.active
        rg = SheetCellRange(ws,"A1:E5")
        self.assertEqual(rg.cells[3:,3:].to_range().coord,"D4:E5")
        self.assertEqual(rg.cells[3:,3:].get_range().coord,"D4:E5") # alias