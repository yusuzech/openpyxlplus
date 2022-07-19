import unittest
import openpyxlplus.utils
from openpyxl import Workbook
from numpy import array as Array
from pandas import Series
from openpyxl import Workbook

class testUtils(unittest.TestCase):
    def test_replicator(self):
        replicator = openpyxlplus.utils.auto_replicate

        case_groups = [
            {
                "values":[
                    1,
                    [1],
                    Series([1]),
                    Array([1])
                ],
                "targets":[
                    [1,2,3,4],
                    Array([1,2,3,4]),
                    Series([1,2,3,4])
                ],
                "expect": [1,1,1,1]
            },
        ]

        for case in case_groups:
            values = case["values"]
            targets = case["targets"]
            expect = case["expect"]
            for target in targets:
                for value in values:
                    self.assertEqual(
                        replicator(value,target),
                        expect,
                        f"Failed for value:{value}, target:{target}"
                    )
        with self.assertRaises(ValueError):
            replicator([1,3],[1,3,5])

    def test_converter(self):
        self.assertEqual(openpyxlplus.utils.coord_to_address((1,1)),"A1")

        self.assertEqual(
            openpyxlplus.utils.boundaries_to_range(1,2,3,4),
            "B1:D3"
        )

        wb = Workbook()
        ws = wb.active

        ws["A1"] = 1
        ws["A2"] = 2
        ws["B1"] = 3
        ws["B2"] = 4

        # clear functions work
        openpyxlplus.utils.clear_range(ws,"A1:B1")
        self.assertIsNone(ws["A1"].value)
        self.assertIsNone(ws["B1"].value)
        self.assertEqual(ws["A2"].value,2)
        self.assertEqual(ws["B2"].value,4)

        openpyxlplus.utils.clear_boundaries(ws,2,1,2,2)
        self.assertIsNone(ws["A2"].value)
        self.assertIsNone(ws["B2"].value)