import unittest
from openpyxl import Workbook,worksheet
import pandas as pd
import numpy as np
from openpyxlplus import writer,cell_range

def is_range_equal_to_array(ws,range_string,array):
    ret = np.array_equal(
        cell_range.SheetCellRange(ws,range_string).cells.get_value(),
        array
    )
    return(ret)

class TestFunctions(unittest.TestCase):
    def test_write_value(self):
        wb = Workbook()
        ws = wb.active
        writer.write_value(1,ws,ws["A1"])
        self.assertEqual(ws["A1"].value,1)

    def test_write_list(self):
        wb = Workbook()
        ws = wb.active
        # down
        writer.write_list([1,2,3],ws,ws["A1"])
        self.assertEqual(
            cell_range.SheetCellRange(ws,"A1:A3").cell_values.flatten().tolist(),
            [1,2,3]
        )
        # up
        # raise Error
        with self.assertRaises(ValueError):
            rg = writer.write_list([1,2,3],ws,ws["A1"],direction="up")
        writer.write_list([1,2,3],ws,ws["A3"],direction="up")
        self.assertEqual(
            cell_range.SheetCellRange(ws,"A1:A3").cell_values.flatten().tolist(),
            [3,2,1]
        )
        # right
        writer.write_list([1,2,3],ws,ws["A1"],direction="right")
        self.assertEqual(
            cell_range.SheetCellRange(ws,"A1:C1").cell_values.flatten().tolist(),
            [1,2,3]
        )

        # left
        with self.assertRaises(ValueError):
            rg = writer.write_list([1,2,3],ws,ws["A1"],direction="left")
        writer.write_list([1,2,3],ws,ws["C1"],direction="left")
        self.assertEqual(
            cell_range.SheetCellRange(ws,"A1:C1").cell_values.flatten().tolist(),
            [3,2,1]
        )

    def test_write_array(self):
        wb = Workbook()
        ws = wb.active
        # write list of list
        writer.write_array([[1,2],[3,4]],ws,ws["A1"])
        self.assertEqual(
            cell_range.SheetCellRange(ws,"A1:B2").cell_values.tolist(),
            [[1,2],[3,4]]
        )

        # fail if dimensions mismatch
        with self.assertRaises(ValueError):
            writer.write_array([[1],[3,4]],ws,ws["A1"])

        # write array
        writer.write_array(np.array([[1,2,3],[2,3,4]]),ws,ws["C3"])
        self.assertEqual(
            cell_range.SheetCellRange(ws,"C3:E4").cell_values.tolist(),
            [[1,2,3],[2,3,4]]
        )

    def test_write_value_merged(self):
        wb = Workbook()
        ws = wb.active
        # form A1 to K1
        writer.write_value_merged(1,ws,right=10)
        self.assertTrue(
            worksheet.cell_range.CellRange("A1:K1") in ws.merged_cells.ranges
        )

        # form A2 to K3
        writer.write_value_merged(1,ws,ws.cell(2,1),down=1,right=10)
        self.assertTrue(
            worksheet.cell_range.CellRange("A2:K3") in ws.merged_cells.ranges
        )

        # negative signs to write up and left
        writer.write_value_merged(1,ws,ws.cell(10,10),up=2,left=2)
        self.assertTrue(
            worksheet.cell_range.CellRange("H8:J10") in ws.merged_cells.ranges
        )
        
        # check if value is correct and centered
        self.assertEqual(ws.cell(8,8).value,1)
        # print(ws.cell(8,8).alignment)
        self.assertEqual(ws.cell(8,8).alignment.horizontal,"center")
        self.assertEqual(ws.cell(8,8).alignment.vertical,"center")


class testWriteDataFrameSingleLevel(unittest.TestCase):
    def setUp(self):
        self.df = pd.DataFrame(
            [[1,2,"a","b"],
            [2,3,"b","c"],
            [3,4,"c","d"]],
            index=["row1","row2","row3"],
            columns=["col1","col2","col3","col4"]
        )
        self.df_values = np.array([
            [None,"col1","col2","col3","col4"],
            ["row1", 1,     2,   "a",    "b" ],
            ["row2", 2,     3,   "b",    "c" ],
            ["row3", 3,     4,   "c",    "d" ]
        ])

        self.df_multilevel = self.df.groupby(["col3","col4"])\
            .agg({"col1":["sum","count"],"col2":["max","min"]})

        self.df_multilevel_values = np.array([
            [None,None,"col1","col1","col2","col2"],
            [None,None,"sum","count", "max", "min"],
            ["a", "b" ,   1,     1,      2,    2  ],
            ["b", "c" ,   2,     1,      3,    3  ],
            ["c", "d" ,   3,     1,      4,    4  ]
        ])
        # initialize class-wise workbook
        self.wb = Workbook()

    def tearDown(self):
        self.wb.close()


    def test_anchor(self):
        ws = self.wb.create_sheet("1")
        # default anchor at A1 works
        writer.write_dataframe(self.df,ws)
        self.assertTrue(is_range_equal_to_array(ws,"A1:E4",self.df_values))
        # anchor works at selected cell
        writer.write_dataframe(self.df,ws,ws["J5"])
        self.assertTrue(is_range_equal_to_array(ws,"J5:N8",self.df_values))

    def test_index_header(self):
        ## Single level
        ws = self.wb.create_sheet("2")
        # index=True, header=True
        writer.write_dataframe(self.df,ws,ws["B2"],index=True,header=True)
        # index=True, header=False
        writer.write_dataframe(self.df,ws,ws["H4"],index=True,header=False)
        self.assertTrue(is_range_equal_to_array(ws,"H4:L6",self.df_values[1:,:]))
        # index=False, header=True
        writer.write_dataframe(self.df,ws,ws["C14"],index=False,header=True)
        self.assertTrue(is_range_equal_to_array(ws,"C14:F17",self.df_values[:,1:]))
        # index=False, header=False
        writer.write_dataframe(self.df,ws,ws["K14"],index=False,header=False)
        self.assertTrue(is_range_equal_to_array(ws,"K14:N16",self.df_values[1:,1:]))

        ## multi level
        ws = self.wb.create_sheet("3")
        # index=True, header=True
        writer.write_dataframe(self.df_multilevel,ws,ws["B2"],index=True,header=True)
        self.assertTrue(is_range_equal_to_array(ws,"B2:G6",self.df_multilevel_values))
        # index=True, header=False
        writer.write_dataframe(self.df_multilevel,ws,ws["H4"],index=True,header=False)
        self.assertTrue(is_range_equal_to_array(ws,"H4:M6",self.df_multilevel_values[2:,:]))
        # # index=False, header=True
        writer.write_dataframe(self.df_multilevel,ws,ws["C14"],index=False,header=True)
        self.assertTrue(is_range_equal_to_array(ws,"C14:F18",self.df_multilevel_values[:,2:]))
        # # index=False, header=False
        writer.write_dataframe(self.df_multilevel,ws,ws["K14"],index=False,header=False)
        self.assertTrue(is_range_equal_to_array(ws,"K14:N16",self.df_multilevel_values[2:,2:]))


# class testUtils(unittest.TestCase):
#     def setUp(self):
#         self.wb = Workbook()

#     def tearDown(self):
#         self.wb.close()

#     def test__merge_consecutive_cells(self):
#         wb = Workbook()
#         ws = wb.active
#         array = [
#             [1,1,2,3],
#             [1,2,2,3],
#             [3,2,3,3]
#         ]
#         # merged cells other than the first one (top left) has value of None.
#         # the styles of merged cells are stored in the first cell (top left)

#         # default on row, center=True
#         rg = writer.write_array(array,ws,cell=ws.cell(1,1))
#         writer._merge_consecutive_cells(rg)

#         self.assertTrue(np.array_equal(
#             rg.cell_values,
#             [
#                 [1,None,2,3],
#                 [1,2,None,3],
#                 [3,2,3,None]
#             ]
#         ))

#         self.assertTrue(np.array_equal(
#             rg.cells.get_style_detail("alignment",["horizontal"]),
#             [
#                 ['center',None,None,None],
#                 [None,'center',None,None],
#                 [None,None,'center',None]
#             ]
#         ))

#         self.assertTrue(np.array_equal(
#             rg.cells.get_style_detail("alignment","vertical"),
#             [
#                 ['center',None,None,None],
#                 [None,'center',None,None],
#                 [None,None,'center',None]
#             ]
#         ))

#         # on column, center=False
#         rg = writer.write_array(array,ws,cell=ws.cell(7,1))
#         writer._merge_consecutive_cells(rg,on="column",center=False)


#         self.assertTrue(np.array_equal(
#             rg.cell_values,
#             [
#                 [1,1,2,3],
#                 [None,2,None,None],
#                 [3,None,3,None]
#             ]
#         ))

#         self.assertTrue(all(
#             rg.cells.get_style_detail("alignment","vertical").flatten() == None
#         ))

#         self.assertTrue(all(
#             rg.cells.get_style_detail("alignment","horizontal").flatten() == None
#         ))

#         wb.close()
        