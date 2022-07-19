import unittest
from pandas import DataFrame
from pandas import MultiIndex
from pandas import Series
from numpy import arange
from numpy import array as Array
from numpy.testing import assert_array_equal
from openpyxl.styles import NamedStyle,Font,Alignment,Border,Side
from openpyxl import Workbook
import openpyxlplus.dataframe
from openpyxlplus.cell_range import SheetCellRange
import numpy as np

class testDataFrame(unittest.TestCase):
    def setUp(self):
        self.df = DataFrame({
            "percent":[0.01,0.5,2],
            "comma":[1000,500,1000000],
            "mixed":["bold_red","centered","all_border"]
        })

        # named styles
        style_percent = NamedStyle("name1",number_format="0.0%")
        style_comma = NamedStyle("name2",number_format="#,##0")
        style_bold_red = NamedStyle("name3",font=Font(bold=True,color="ff0000"))
        style_centered = NamedStyle(
            "name4",
            alignment=Alignment(horizontal="center",vertical="center")
        )
        style_all_border = NamedStyle(
            "name5",
            border=Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        )
        # style data frame
        self.style_array = DataFrame({
            "percent":[style_percent] * 3,
            "comma":[style_comma] * 3,
            "mixed":[style_bold_red,style_centered,style_all_border]
        })
        # style data frame with wrong size
        self.wrong_style_array = arange(12).reshape(3,4)
        # initialize class-wise workbook
        self.wb = Workbook()
        self.ws = self.wb.active

    def tearDown(self):
        self.wb.close()

    def test_offset_headers(self):
        wb = Workbook()
        ws = wb.active
        testdf = DataFrame({
            "test1":[1,2,3],
            "test2":[2,3,4]
        })

        style_bold = NamedStyle("bold",font=Font(bold=True))
        style_red = NamedStyle("red",font=Font(color="ff0000"))
        # # won't allow to use styles with different shape
        with self.assertRaises(ValueError):
            openpyxlplus.dataframe.df_to_excel(
                self.df,self.ws,
                body_styles = self.wrong_style_array
            )
        # offset works
        openpyxlplus.dataframe.df_to_excel(testdf,ws,row=2,col=2)
        # skip first row and first column
        self.assertIsNone(ws['A1'].value)
        self.assertIsNone(ws['A2'].value)
        self.assertIsNone(ws['B1'].value)
        # starting at B2
        self.assertEqual(ws['B2'].value,"test1")
        self.assertEqual(ws['B3'].value,1)
        self.assertEqual(ws['B5'].value,3)

        # no header works
        openpyxlplus.dataframe.df_to_excel(testdf,ws,row=2,col=2,headers=False)
        self.assertEqual(ws['B2'].value,1)

        openpyxlplus.dataframe.df_to_excel(testdf,ws,row=2,col=2,headers=None)
        self.assertEqual(ws['B2'].value,1)

        # assign headers work
        openpyxlplus.dataframe.df_to_excel(testdf,ws,row=1,col=1,headers=["x","y"])
        self.assertEqual(ws['A1'].value,"x")
        self.assertEqual(ws['B1'].value,"y")

        # header style works
        # single
        openpyxlplus.dataframe.df_to_excel(
            testdf,ws,row=1,col=1,header_styles=style_bold
        )
        self.assertEqual(ws['B1'].font.b,True)

        #list multiple
        openpyxlplus.dataframe.df_to_excel(
            testdf,ws,row=1,col=1,header_styles=[style_red]
        )
        self.assertEqual(ws['B1'].font.color.rgb,"00ff0000")

        # list works
        openpyxlplus.dataframe.df_to_excel(
            testdf,ws,row=1,col=1,header_styles=[style_bold,style_red]
        )
        self.assertEqual(ws['A1'].font.b,True)
        self.assertEqual(ws['B1'].font.color.rgb,"00ff0000")
        # array works
        openpyxlplus.dataframe.df_to_excel(
            testdf,ws,row=1,col=1,header_styles=Array([style_bold,style_red])
        )
        self.assertEqual(ws['A1'].font.b,True)
        self.assertEqual(ws['B1'].font.color.rgb,"00ff0000")
        # pandas series works
        openpyxlplus.dataframe.df_to_excel(
            testdf,ws,row=1,col=1,header_styles=Series([style_bold,style_red])
        )
        self.assertEqual(ws['A1'].font.b,True)
        self.assertEqual(ws['B1'].font.color.rgb,"00ff0000")
        wb.close()

    def test_body(self):
        openpyxlplus.dataframe.df_to_excel(self.df,self.ws,body_styles=self.style_array)
        # column headers added
        self.assertEqual(self.ws['A1'].value,"percent")
        self.assertEqual(self.ws['C1'].value,"mixed")
        # types are correct
        self.assertEqual(self.ws['A2'].number_format,"0.0%")
        self.assertEqual(self.ws['B2'].number_format,"#,##0")

        self.assertEqual(self.ws['C2'].font.color.rgb,"00ff0000")
        self.assertEqual(self.ws['C2'].font.bold,True)

        self.assertEqual(self.ws['C3'].alignment.horizontal,"center")
        self.assertEqual(self.ws['C3'].alignment.vertical,"center")

        for side in ["left","top","right","bottom"]:
            self.assertEqual(getattr(self.ws['C4'].border,side).style,"thin")



class testMultiHeader(unittest.TestCase):
    def setUp(self):
        self.df=DataFrame({'a':[1,2,3],'b':[4,5,6],"c":[7,8,9],"d":[10,11,12]})
        columns=[('c','a'),('c','b'),('d','a'),('d','b')]
        self.df.columns=MultiIndex.from_tuples(columns)

        self.style_center = NamedStyle(
            "Style2",
            alignment=Alignment(horizontal="center",vertical="center")
        )

    def test_headers(self):
        wb = Workbook()
        ws = wb.active
        with self.assertRaises(ValueError):
            # doesn't work if shaped is incorrect
            openpyxlplus.dataframe.df_to_excel(
                self.df,
                ws,
                row=2,
                header_styles= Array([1,2])
            )

        openpyxlplus.dataframe.df_to_excel(
            self.df,
            ws,
            row=2,
            header_styles= [
                [self.style_center] * 4,
                [None] * 4
            ]
        )
        a = []
        for row in range(2,7):
            l = []
            for col in range(1,5):
                l.append(ws.cell(row,col).value)
            a.append(l)
        # data written correctly
        assert_array_equal(
            Array(a),
            Array([['c',None,'d',None],
                    ['a','b','a','b'],
                    [ 1,  4,  7, 10],
                    [ 2,  5,  8, 11],
                    [ 3,  6,  9, 12]])
        )
        # cells are merged correctly
        self.assertEqual(len(ws.merged_cells.ranges),2)
        # header merge style is applied correctly
        self.assertEqual(ws.cell(2,1).alignment.vertical,"center")
        self.assertEqual(ws.cell(2,1).alignment.horizontal,"center")


def is_range_equal_to_array(ws,range_string,array):
    ret = np.array_equal(
        SheetCellRange(ws,range_string).cells.get_value(),
        array
    )
    return(ret)

class testDataFrameMinimal(unittest.TestCase):
    def setUp(self):
        self.df = DataFrame({
            "percent":[0.01,0.5,2],
            "comma":[1000,500,1000000],
            "mixed":["bold_red","centered","all_border"]
        })
        # initialize class-wise workbook
        self.wb = Workbook()

    def tearDown(self):
        self.wb.close()

    def test_anchor(self):
        array_value = np.array(
            [['percent', 'comma', 'mixed'],
            [0.01, 1000, 'bold_red'],
            [0.5, 500, 'centered'],
            [2.0, 1000000, 'all_border']],dtype='O'
        )
        # at 2,2
        self.wb.create_sheet("sheet1")
        ws = self.wb["sheet1"]
        openpyxlplus.dataframe.df_to_sheet_minimal(
            df = self.df,
            ws = ws,
            row_anchor=2,
            col_anchor=2,
            index=False,
            header=True
        )
        self.assertTrue(is_range_equal_to_array(ws,"B2:D5",array_value))

        # at 10,3
        self.wb.create_sheet("sheet2")
        ws = self.wb["sheet2"]
        openpyxlplus.dataframe.df_to_sheet_minimal(
            df = self.df,
            ws = ws,
            row_anchor=11,
            col_anchor=3,
            index=False,
            header=True
        ) 
        self.assertTrue(is_range_equal_to_array(ws,"C11:E14",array_value))

    def test_index_header(self):
        # index and header
        array_value = np.array(
            [[None,'percent', 'comma', 'mixed'],
            [None,None,None,None],
            [0,0.01, 1000, 'bold_red'],
            [1,0.5, 500, 'centered'],
            [2,2.0, 1000000, 'all_border']]
        )
        self.wb.create_sheet("sheet3")
        ws = self.wb["sheet3"]
        openpyxlplus.dataframe.df_to_sheet_minimal(
            df = self.df,
            ws = ws,
            index=True,
            header=True
        ) 
        self.assertTrue(is_range_equal_to_array(ws,"A1:D5",array_value))
        # index
        array_value = np.array(
            [[None,"percent","comma","mixed"],
            [0,0.01, 1000, 'bold_red'],
            [1,0.5, 500, 'centered'],
            [2,2.0, 1000000, 'all_border']],dtype="O"
        )
        self.wb.create_sheet("sheet3")
        ws = self.wb["sheet3"]
        openpyxlplus.dataframe.df_to_sheet_minimal(
            df = self.df,
            ws = ws,
            index=True,
            header=False
        )
        self.assertTrue(is_range_equal_to_array(ws,"A1:D4",array_value))
        # skip: header tested at test_anchor


class testDataFrameSimple(unittest.TestCase):
    def setUp(self):
        self.df = DataFrame({
            "percent":[0.01,0.5,2],
            "comma":[1000,500,1000000],
            "mixed":["bold_red","centered","all_border"]
        })
        self.style_percent = NamedStyle("name1",number_format="0.0%")
        self.style_comma = NamedStyle("name2",number_format="#,##0")
        # initialize class-wise workbook
        self.wb = Workbook()

    def tearDown(self):
        self.wb.close()

    def test_default_index_header(self):
        # index = False, header = True ----
        self.wb.create_sheet("sheet1")
        ws = self.wb["sheet1"]

        openpyxlplus.dataframe.df_to_sheet_simple(
            df = self.df,
            ws = ws,
            row_anchor=2,
            col_anchor=2,
            index=False,
            header=True
        )
        self.assertEqual(ws['B2'].font.bold,True)
        self.assertEqual(ws['C2'].font.bold,True)
        self.assertEqual(ws['D2'].font.bold,True)

        # index = True, header = True ----
        self.wb.create_sheet("sheet2")
        ws = self.wb["sheet2"]

        openpyxlplus.dataframe.df_to_sheet_simple(
            df = self.df,
            ws = ws,
            index=True,
            header=True
        )

        # first row is bold
        self.assertEqual(ws['A1'].font.bold,True)
        self.assertEqual(ws['B1'].font.bold,True)
        self.assertEqual(ws['C1'].font.bold,True)
        self.assertEqual(ws['D1'].font.bold,True)
        # first column is bold
        self.assertEqual(ws['A1'].font.bold,True)
        self.assertEqual(ws['A2'].font.bold,True)
        self.assertEqual(ws['A3'].font.bold,True)
        self.assertEqual(ws['A4'].font.bold,True)
        self.assertEqual(ws['A4'].font.bold,True)

    def test_styles(self):
        # change style by column
        self.wb.create_sheet("sheet3")
        ws = self.wb["sheet3"]
        openpyxlplus.dataframe.df_to_sheet_simple(
            df = self.df,
            ws = ws,
            index=False,
            header=True,
            styles = [self.style_percent,self.style_comma],
            style_axis=1
        )

        self.assertEqual(ws["A2"].number_format,"0.0%")
        self.assertEqual(ws["B2"].number_format,"#,##0")
        self.assertEqual(ws["C2"].number_format,"0.0%")

        # change style by row
        self.wb.create_sheet("sheet4")
        ws = self.wb["sheet4"]
        openpyxlplus.dataframe.df_to_sheet_simple(
            df = self.df,
            ws = ws,
            index=False,
            header=True,
            styles = [self.style_percent,self.style_comma],
            style_axis=0
        )
        self.assertEqual(ws["A2"].number_format,"0.0%")
        self.assertEqual(ws["A3"].number_format,"#,##0")
        self.assertEqual(ws["A4"].number_format,"0.0%")

    def test_returned_table_range(self):
        # with index and header
        self.wb.create_sheet("sheet5")
        ws = self.wb["sheet5"]
        table_range5 = openpyxlplus.dataframe.df_to_sheet_simple(
            df = self.df,
            ws = ws,
            index=True,
            header=True
        )
        self.assertEqual(table_range5.header.coord,"B1:D2")
        self.assertEqual(table_range5.index.coord,"A3:A5")
        self.assertEqual(table_range5.body.coord,"B3:D5")
        # with index
        self.wb.create_sheet("sheet6")
        ws = self.wb["sheet6"]
        table_range5 = openpyxlplus.dataframe.df_to_sheet_simple(
            df = self.df,
            ws = ws,
            index=True,
            header=False
        )
        self.assertEqual(table_range5.header.coord,"B1:D1")
        self.assertEqual(table_range5.index.coord,"A2:A4")
        self.assertEqual(table_range5.body.coord,"B2:D4")
        # with header
        self.wb.create_sheet("sheet7")
        ws = self.wb["sheet7"]
        table_range5 = openpyxlplus.dataframe.df_to_sheet_simple(
            df = self.df,
            ws = ws,
            index=False,
            header=True
        )
        self.assertEqual(table_range5.header.coord,"A1:C1")
        self.assertIsNone(table_range5.index)
        self.assertEqual(table_range5.body.coord,"A2:C4")