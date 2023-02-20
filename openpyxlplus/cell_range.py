from openpyxl.worksheet.cell_range import CellRange
from . import utils
from openpyxl.styles import Border,Side,Alignment
import numpy as np
from copy import copy
import re
import openpyxl.utils.cell as converter

def getattr_copy(obj,name,*args):
    if len(args) > 0:
        return(copy(getattr(obj,name,args[0])))
    else:
        return(copy(getattr(obj,name)))

v_getattr = np.vectorize(getattr_copy,otypes="O")
v_setattr = np.vectorize(setattr,otypes="O")

class SheetCellRange(CellRange):
    """
    Inheritate from openpyxl.worksheet.cell_range.CellRange. This class allow 
        additional operations on a given cell range such as:
    - Convert SheetCellRange to Cells which is a wrapper on numpy.ndarray.
    - Get cell or cells with coordinate(s).

    """
    def __init__(self,ws, range_string=None, min_col=None, min_row=None,
                 max_col=None, max_row=None, title=None):
        self.ws = ws
        super().__init__(range_string=range_string, min_col=min_col, 
            min_row=min_row, max_col=max_col, max_row=max_row, title=title)
    
    @property
    def cells(self):
        """
        Return cells in range as Cells object. This needs to be a property 
            because operations like shrink,expand,shift could change cell range.
        """
        cells = []
        for row in self.rows:
            l = []
            for cell in row:
                l.append(self.ws.cell(cell[0],cell[1]))
            cells.append(l)
        return(Cells(cells))

    def get_cell(self,coordinate,relative=True):
        """
        Parameters:
        -------------
        coordinate: coordinate of cell in sheet (row:int,column:int)
        relative: Default to True, uses coordinate relative to top left 
            coordinate of this cell range. Additionally, when relative = True, 
            the coordinate start from 0. When relative = False, uses absolute 
            coordinate in the sheet, which index starts from 0.

        Return:
        ---------
        openpyxl.cell.cell.Cell object
        """
        if relative:
            ret = self.ws.cell(
                coordinate[0] + self.min_row,
                coordinate[1] + self.min_col
            )
        else:
            ret = self.ws.cell(coordinate[0],coordinate[1])
        return(ret)
    
    def get_cells(self,coordinates,relative = True):
        """
        subset Cells with coordinates stored in list or numpy.ndarray. Result
        will the same shape as shape of coordinates

        Parameters:
        -------------
        coordinates: list/tuple or numpy.ndarray which contains coordinate tuple 
         with format of (row:int,column:int)
        relative: whether the coordinates are relative to range. See details in
            .get_cells method

        Return:
        ---------
        Cells object
        """
        try:
            coordinates[0][0] # test depth of coordinates
            ret = []
            for row in coordinates:
                l = []
                for cell in row:
                    l.append(self.get_cell(cell,relative))
                ret.append(l)
            ret = Cells(ret)
        except:
            ret = Cells([self.get_cell(x,relative) for x in coordinates])
        return(ret)

    def get_subset(
        self,
        slice_string=None,
        min_col_idx=None,
        min_row_idx=None,
        max_col_idx=None,
        max_row_idx=None
    ):
        """
        Subsetting range by either a slicing string(numpy style) or providing
        min/max of column/row index.

        Parameters:
        -------------
        sliece_string: a numpy style subsetting string such as "1:,:-2"
        min_col_idx, min_row_idx, max_col_ids, max_row_idx: min/max index for 
            rows and column. index starting from 0.
        """
        if slice_string is None:
            row_min = min_row_idx if min_row_idx else ""
            row_max = max_row_idx if max_row_idx else ""
            col_min = min_col_idx if min_col_idx else ""
            col_max =  max_col_idx if max_col_idx else ""
            slice_string = f"{row_min}:{row_max},{col_min}:{col_max}"

        slice_string = re.sub("[^0-9-,:]","",slice_string) # sanitize input
        ret = eval(f"self.cells[{slice_string}]").to_range()
        return(ret)

    def clear(self,value=True,formatting=True):
        """
        Clear value/formatting in range

        Parameters:
        -------------
        value: default True, clear values
        formatting: default True, clear formatting
        """
        if value:
            self.write(None,keep_style=True)

        if formatting:
            self.cells.set_style("style","Normal")
        return(self)

    def merge_cells(self):
        """
        Merge all cells in range
        """
        self.ws.merge_cells(range_string=self.coord)
        return(self)

    def unmerge_cells(self):
        """
        Unmerge all Cells in range
        """
        self.ws.unmerge_cells(range_string=self.coord)
        return(self)

    def write(self,data,keep_style=True):
        """
        Alias to self.set_value. Write data to this range.

        Parameters:
        -------------
        data: scaler, list or numpy array. Note that the shape of data should 
            match the shape of range to ensure writing correctly.
        keep_style: whether to keep original style
        """
        self.cells.set_value(data,keep_style=keep_style)
        return(self)

    def add_border(self,side = None,left=True,right=True,top=True,bottom=True):
        """
        Add outside border to cells.

        Parameters:
        -------------
        side: Side object. If None, use Side(style="thin")
        left: whether to add border to left
        right: whether to add border to right
        head: whether to add border to top
        tail: whether to add border to tail
        """
        self.cells.add_border(
            side=side,left=left,right=right,top=top,bottom=bottom
        )
        return(self)

    def merge_consecutive_cells(self,on="row",center=True):
        """
        Merge consecutive cells by rows or by columns

        Parameters:
        -------------
        on: "row" or "column". "row" to merge horizontally (same columns are merged);
            "column" to merge vertically (different columns are merged)
        center: True to center merged cells.
        """
        if on == "row":
            temp = self.cells
        elif on == "column":
            temp = self.cells.transpose()
        else:
            raise Exception(f"{on} not supported.")
        groups = []
        for row in temp:
            g = []
            for cell in row:
                if len(g) == 0:
                    g.append(cell)
                else:
                    if cell.value == g[-1].value:
                        g.append(cell)
                    else:
                        groups.append(g)
                        g = [cell]
            groups.append(g)

        for g in groups:
            if len(g) > 1:
                rg_temp = Cells(g).to_range()
                if center:
                    rg_temp.cells.modify_style(
                        "alignment",
                        Alignment(horizontal='center',vertical="center")
                    )
                rg_temp.merge_cells()
        return(self)

    @property
    def cell_values(self):
        """
        Display a copy of values in range as numpy array.
        """
        return(self.cells.get_value())

    # def autofit_width(self,max_width=None,multiplier=1.2):
    #     """
    #     Automatically fit column width with current cell range by detecting number 
    #         of characters in each cell. 

    #     Parameters:
    #     max_width: maximum column width
    #     multiplier: increase or reduce column width globally in current range. 
    #         This value may be adjusted for better view.

    #     Return:
    #     self
    #     """     
    #     max_length = ((max_width if max_width else 999) - 2)/multiplier

    #     for col_coord in self.cols:
    #         col_max_length = max_length
    #         lengths = []

    #         for cell_coord in col_coord:
    #             cell = self.ws.cell(*cell_coord)
    #             if cell.value:
    #                 lengths.append(len(str(cell.value)))
    #             else:
    #                 lengths.append(0)
    #         # set column width
    #         cell_max_length = max(lengths)
    #         length = col_max_length if cell_max_length > col_max_length else cell_max_length
    #         self.ws.column_dimensions[converter.get_column_letter(cell.column)].width = \
    #             int((length + 2) * multiplier)
    #     return(self)

    def autosize(
        self,
        adjust_width=True,
        adjust_height=True,
        **kwargs
    ):
        """
        Automatically adjust sheet width and height using given range. Based on
        number of characters in each cell and width/height factor.

        if adjust_height = True:
            height will be determined by calc_value_shape function which counts 
            number of newline (\\n) in text (which is not desired in most case).
        if adjust_height = False:
            height will be determined automatically by application.

        Check openpyxlplus.utils.calc_value_size for more details.

        Parameters:
        -------------
        adjust_width: True/False. If False, won't adjust width
        adjust_height: True/False. If False, won't adjust height

        Below are kwargs and their default value:
        min_width: 1.43
        min_height:  13.2
        max_width: 255
        max_height: 409
        width_factor: 0.11
        height_factor: 1.35
        max_ndigits: 2 
        wrap_text: False

        """
        cells = self.cells
        cell_values = cells.get_value()
        fontsizes = cells.get_style_detail("font","size")

        heights = np.zeros_like(cell_values)
        widths = np.zeros_like(cell_values)
        for ind,v in np.ndenumerate(cell_values):
            ftsize = fontsizes[ind] 
            h,w = utils.calc_value_size(v,fontsize=ftsize,**kwargs)
            heights[ind] = h
            widths[ind] = w
        # print(heights)
        # print(widths)

        if adjust_width:
            column_widths = widths.max(axis=0)
            column_letters = [converter.get_column_letter(x.column) for x in cells[0,:]]
            for column_letter,column_width in zip(column_letters,column_widths):
                self.ws.column_dimensions[column_letter].width = column_width
        if adjust_height:
            row_heights = heights.max(axis=1)
            row_numbers = [x.row for x in cells[:,0]]
            for row_number,row_height in zip(row_numbers,row_heights):
                self.ws.row_dimensions[row_number].height = row_height

        return(self)

    def show_in_excel(self,temp_name="temp.xlsx"):
        """
        Show this range in excel application.

        Parameters:
        -------------
        temp_name: temporary file name. Default to "temp.xlsx"
        """
        _ , workbook = utils.open_workbook(self.ws.parent,temp_name=temp_name
            ,prompt=False)

        ws = workbook.Sheets(self.ws.title)
        ws.Select() # make this sheet active
        ws.Range(self.coord).Select()
        print("Remember to close the workbook manually.")
        # Wait before closing it
        # _ = input("Press enter to close Excel")
        # excel_app.Application.Quit()

    # method alias
    show = show_in_excel

class SheetTableRange(SheetCellRange):
    """
    Create a CellRange that represents a table.

    Additional attributes:
    header: a Cellrange object for table header
    index: a Cellrange object for table index
    body: a Cellrange object for table body (not index or header)
    
    """
    def __init__(self,ws,range_string=None, min_col=None, min_row=None,
                max_col=None, max_row=None, title=None,n_index=0,n_header=1):
        """
        Parameters:
        -------------
        n_index: integer, how many columns from left are index
        n_header: integer, how many rows from top are header 
        """
        super().__init__(ws=ws,range_string=range_string, min_col=min_col, 
            min_row=min_row, max_col=max_col, max_row=max_row, title=title)
        self.n_index = n_index
        self.n_header = n_header

        
    @property
    def header(self):
        if self.n_header == 0:
            return(None)
        else:
            ret = SheetCellRange(
                self.ws,
                min_col = self.min_col + self.n_index,
                min_row = self.min_row,
                max_col = self.max_col,
                max_row = self.min_row + self.n_header-1,
            )
            return(ret)
    @property
    def index(self):
        if self.n_index == 0:
            return(None)
        else:
            ret = SheetCellRange(
                self.ws,
                min_col = self.min_col,
                min_row = self.min_row + self.n_header,
                max_col = self.min_col + self.n_index - 1,
                max_row = self.max_row,
            )
            return(ret)

    @property
    def body(self):
        ret = SheetCellRange(
            self.ws,
            min_col = self.min_col + self.n_index,
            min_row = self.min_row + self.n_header,
            max_col = self.max_col,
            max_row = self.max_row,
        )
        return(ret)

    @property
    def top_left_corner(self):
        """
        When index and header are both enabled, it creates a blank area at top 
        left corner of the table.

        Note: openpyxl.utils.dataframe.dataframe_to_rows adds index name when
        index = True. This may cause confusion sometimes, so it is advised to
        set index = False.
        """
        if self.n_index == 0 or self.n_header == 0:
            return(None)
        else:
            ret = SheetCellRange(
                self.ws,
                min_col = self.min_col,
                min_row = self.min_row,
                max_col = self.min_col + self.n_index - 1,
                max_row = self.min_row + self.n_header - 1,
            )
            return(ret)

class Cells(np.ndarray):
    """
    Wrapper class on numpy.ndarray. Allows following operations:

    - Read values, attributes from the range or subset of Cells
    - Apply style to range or subset of range
    - Write or change values in range
    - Add outside border to the range
    - merge/unmerge cells in range

    returned values preserve the shape of array.
    """
    def __new__(cls,array):
        array = np.array(array).view(cls)
        return(array)


    @staticmethod
    def from_range(ws, range_string=None, min_col=None, min_row=None,
        max_col=None, max_row=None, title=None):
        """
        Construct Cells from range string or min/max col/row.
        """
        ret = SheetCellRange(ws, range_string=range_string, min_col=min_col, 
            min_row=min_row, max_col=max_col, max_row=max_row, title=title).cells
        return(ret)

    def get_style(self,style_name):
        """
        Get a copy of cells's style as numpy.ndarray
        """
        ret = v_getattr(self,style_name).view(np.ndarray).copy()
        return(ret)

    def get_style_detail(self,style_name,detail):
        """
        Get detail in style

        Parameters:
        -------------
        style_name: name of cell style
        details: attribute name of the detail, if detail is nested, provide a 
            list of attribute names or style names separated by dots
            e.g. self.get_style_detail("Border",["left","style"]) 
            or self.get_style_detail("Border","left.style") 
        """
        ret = self.get_style(style_name)
        if type(detail) == str:
            if "." in detail:
                detail = detail.split(".")
            else:
                detail = [detail]
        for x in detail:
            ret = v_getattr(ret,x)
        ret = ret.copy()
        return(ret)

    def set_style(self,style_name,style,overwrite=True):
        """
        Set cells' style to value(s), overwrites original style. returns self
        
        Parameters:
        -------------
        style_name: style name to change
        style: desired style. Note that the shape of data should 
            match the shape of range to ensure writing correctly.
        overwrite: default True. if set to False, this function is equivalent to
            self.modify_style
        """
        if overwrite:
            v_setattr(self,style_name,style)
        else:
            self.modify_style(style_name,style)
        return(self)

    def modify_style(self,style_name,style):
        """
        Modify cells' style with given style(s). This method adds style(s) to 
        original one instead of completely overwriting it.
        """
        v_setattr(self,style_name,v_getattr(self,style_name)+style)
        return(self)


    def get_value(self):
        """
        Get a copy of cells' values
        """
        ret = v_getattr(self,"value").view(np.ndarray).copy().astype('O')
        return(ret)

    def set_value(self,value,keep_style = True):
        """
        Set cells values with given value(s)

        Parameters:
        -----------
        value: value to write to cells.Note that the shape of data should 
            match the shape of range to ensure writing correctly.
        keep_style: whether to preserve original style. If set to False, clear
            original style.
        """
        # # 0.5.0 keep style manually
        # if keep_style:
        #     original_style = self.get_style("_style")
        #     v_setattr(self,"value",value)
        #     self.set_style("_style",original_style)
        # else:
        #     v_setattr(self,"value",value)

        # 0.5.1 keep style automatically
        if keep_style:
            v_setattr(self,"value",value)
        else:
            v_setattr(self,"value",value)
            self.set_style("style","Normal")
        return(self)

    def get_range(self):
        """
        Convert Cells to SheetCellRange using top left and bottom right cell 
            coordinates.
        """
        try:
            top_left = self[0][0]
            bottom_right = self[-1][-1]
        except Exception:
            top_left = self[0]
            bottom_right = self[-1]
        ws = top_left.parent
        ret = SheetCellRange(
            ws,f"{top_left.coordinate}:{bottom_right.coordinate}"
        )
        return(ret)
    
    to_range = get_range

    def side(self,direction,n):
        """
        Get n columns or rows from given direction to center.
        """
        if direction == "head":
            ret = self[:n,:]
        elif direction == "tail":
            ret = self[-n:,:]
        elif direction == "left":
            ret = self[:,:n]
        elif direction == "right":
            ret = self[:,-n:]
        else:
            raise ValueError(f"direcion:'{direction}'' is not " +\
            "one of 'head', 'tail', 'left' or 'right'")
        return(ret)
    
    def head(self,n=1):
        return(self.side("head",n))

    # method alias
    top = head

    def tail(self,n=1):
        return(self.side("tail",n))

    # method alias
    bottom = tail

    def left(self,n=1):
        return(self.side("left",n))
        
    def right(self,n=1):
        return(self.side("right",n))

    def add_border(self,side = None,left=True,right=True,top=True,bottom=True):
        """
        Add outside border to cells.

        Parameters:
        -------------
        side: Side object. If None, use Side(style="thin")
        left: whether to add border to left
        right: whether to add border to right
        head: whether to add border to top
        tail: whether to add border to tail
        """
        if side is None:
            side = Side(style="thin")
        if left:
            self.left().modify_style("border",Border(left=side))
        if right:
            self.right().modify_style("border",Border(right=side))
        if top:
            self.head().modify_style("border",Border(top=side))
        if bottom:
            self.tail().modify_style("border",Border(bottom=side))
        return(self)