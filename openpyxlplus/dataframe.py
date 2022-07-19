"""
Please use writer and cell_range module instead.
"""
from openpyxl.styles import Font, NamedStyle
from pandas import Series,DataFrame, MultiIndex
from numpy import array as Array
from numpy import newaxis
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxlplus.utils import auto_replicate
from openpyxlplus.cell_range import TableRange
from functools import reduce
from itertools import cycle

def df_to_excel(
    df,
    worksheet,
    row=1,col=1,
    headers = True,
    header_styles = None,
    index = False,
    body_styles = None,
):
    """
    Write dataframe to excel worksheet with offset and styles.
    NOT UNDER Maintenance.
    
    Parameters:
    df: pandas dataframe
    worksheet: openpyxl worksheet
    row: starting row
    col: starting column
    headers: True, False, or a list/np array with the same length of columns in 
        df. True uses default columns, None won't create columns, list can
        customize columns.
    header_styles: None or a NamedStyle or a numpy array of NamedStyle equal to the 
        shape of headers. If headers is multiIndex, provide a numpy array 
        with equal dimension.
    index: Whether to write index to excel, default to False, True not implemented.
    body_styles: numpy array or dataframe of styles with the same dimension as 
        df.shape.
        
    """
    # header and header styles
    if type(headers) is bool:
        if headers:
            headers = df.columns.values
        else:
            headers = None
    # if header exist
    if headers is not None:
        if isinstance(df.columns,MultiIndex):
            levels = len(headers[0])
            row_anchor = levels
            # if none, create empty header styles
            if header_styles is None:
                header_styles = Array([None] * (levels * len(headers)))[newaxis]
                header_styles = header_styles.reshape((levels,len(headers)))
            # if NamedStyle, apply named style to all header cells
            if isinstance(header_styles,NamedStyle):
                header_styles = Array([header_styles] * (levels * len(headers)))[newaxis]
                header_styles = header_styles.reshape((levels,len(headers)))               

            # convert header_styles to array
            header_styles = Array(header_styles)
            if header_styles.shape != (levels,len(headers)):
                raise ValueError(
                f"Style array shape {header_styles.shape} "
                f"is not equal to headers's shape {(levels,len(headers))}"
            )
            # transpose array
            new_headers = []
            for i in range(levels):
                new_headers.append(list(map(lambda x: x[i],df.columns.values)))
            # merge columns if there have the same value and are connected rowwise
            for rowindex,header_row in enumerate(new_headers):
                previous_value = None
                col_merge_begin = -1
                for colindex,value in enumerate(header_row):
                    if value != previous_value:
                        cell = worksheet.cell(
                            row + rowindex,
                            col + colindex,
                            value = value
                        )
                        if colindex-1 != col_merge_begin:
                            worksheet.merge_cells(
                                start_row=row + rowindex, 
                                start_column=col + col_merge_begin, 
                                end_row=row + rowindex, 
                                end_column=col + colindex -1
                            )
                        col_merge_begin = colindex
                        previous_value = value
                        # add style
                        if header_styles[rowindex,colindex] is not None:
                            cell.style = header_styles[rowindex,colindex]
                    else:
                        pass
                # special case check should the last group be merged
                if colindex != col_merge_begin:
                    worksheet.merge_cells(
                                    start_row=row + rowindex, 
                                    start_column=col + col_merge_begin, 
                                    end_row=row + rowindex, 
                                    end_column=col + colindex
                                )

        else:
            header_styles = auto_replicate(header_styles,headers)
            # write headers
            for colindex,value in enumerate(headers):
                cell = worksheet.cell(row,col + colindex,value = value)
                if header_styles[colindex] is not None:
                    cell.style = header_styles[colindex]
            row_anchor = 1
    # if header doesn't exist
    else:
        row_anchor = 0


    # body styles
    if body_styles is None:
        body_styles = Array([None] * (df.shape[0] * df.shape[1]))\
            .reshape(df.shape)
    elif type(body_styles) == list:
        body_styles = Array(body_styles)
    elif isinstance(body_styles,DataFrame):
        body_styles = body_styles.values
    if body_styles.shape != df.shape:
        raise ValueError((
            f"body_styles' shape({body_styles.shape}) "
            f"and dataframe's shape({df.shape}) are not identical"
        ))
    

    for rowindex,rowl in enumerate(dataframe_to_rows(df,index=False,header=False)):
        for colindex,value in enumerate(rowl):
            cell = worksheet.cell(
                row + rowindex + row_anchor,col + colindex,
                value = value
            )
            cell_style = body_styles[(rowindex,colindex)]
            if cell_style is not None:
                cell.style = body_styles[(rowindex,colindex)]
    return(True)


def df_to_sheet_minimal(
    df,
    ws,
    index = False,
    header = True,
    row_anchor = 1,
    col_anchor = 1
):
    """
    Write a dataframe to excel worksheet. 

    Parameters:
    df: dataframe to write
    ws: worksheet to write to
    index: whether to write index. True or False. Please note that there is a
        behavioral difference because of how openpyxl.utils.dataframe_to_rows 
        works. When setting index=True, the second row will be empty except the 
        index.
    header: whether to write header. True or False
    row_anchor: row number for top left corner of dataframe. minimum value is 1.
    col_anchor: column number for top left corner of dataframe. minimum value is 1.
    
    Return:
    openpyxlplus.cell_range.TableRange object. It allows easier modification of
        values and styles after the data is written to the sheet
    """
    index_nlevels = df.index.nlevels if index else 0
    header_nlevels = df.columns.nlevels if header else 0

    for i,row in enumerate(dataframe_to_rows(df,index=index,header=header)):
        for j, cell in enumerate(row):
            row_num = row_anchor + i
            col_num = col_anchor + j
            ws.cell(row_num,col_num,value=cell)
    # table range:
    min_col = col_anchor
    min_row = row_anchor
    max_col = min_col + j
    max_row = min_row + i
    n_header = header_nlevels + int(index) if index else header_nlevels
    table_range = TableRange(ws,min_col=min_col,min_row=min_row,max_col=max_col,
        max_row=max_row,n_index=index_nlevels,n_header=n_header)
    return(table_range)

def df_to_sheet_simple(
    df,
    ws,
    index = False,
    header = True,
    row_anchor = 1,
    col_anchor = 1,
    index_style = "default",
    header_style = "default",
    styles = None,
    style_axis = 1
):
    """
    Write a dataframe to excel worksheet with option to set header, index 
        column/row style.

    Parameters:
    df: dataframe to write
    ws: worksheet to write to
    index: whether to write index. True or False. Please note that there is a
        behavioral difference because of how openpyxl.utils.dataframe_to_rows 
        works. When setting index=True, the second row will be empty except the 
        index.
    header: whether to write header. True or False
    row_anchor: row number for top left corner of dataframe. minimum value is 1.
    col_anchor: column number for top left corner of dataframe. minimum value is 1.
    index_style: None, "default" or a NamedStyle. When default, bold font will 
        be used.
    header_style: None, "default" or a NamedStyle. When default, bold font will 
        be used.
    styles: list of NamedStyle or None to apply to dataframe rows or columns. 
        Will cycle styles if length is smaller than length or row or column.
    style_axis: Which axis to apply styles to. 0 for rows, 1 for columns.
    
    Return:
    openpyxlplus.cell_range.TableRange object. It allows easier modification of
        values and styles after the data is written to the sheet
    """
    default_style = NamedStyle(name="Normal")
    if styles:
        for i,s in enumerate(styles):
            if s is None:
                styles[i] = default_style
        styles_cycle = cycle(styles)
    else:
        styles_cycle = cycle([default_style])

    index_nlevels = df.index.nlevels if index else 0
    header_nlevels = df.columns.nlevels if header else 0

    j = 0
    for i,row in enumerate(dataframe_to_rows(df,index=index,header=header)):
        # if apply style by row
        if i >= header_nlevels and j >= index_nlevels:
            if style_axis == 0:
                current_style = next(styles_cycle)
        for j, cell in enumerate(row):
            row_num = row_anchor + i
            col_num = col_anchor + j
            ws.cell(row_num,col_num,value=cell)

            # set header style
            if header:
                if i < header_nlevels:
                    if header_style == "default":
                        ws.cell(row_num,col_num).font = Font(bold=True)
                    elif header_style is None:
                        ws.cell(row_num,col_num).style = default_style
                    else:
                        ws.cell(row_num,col_num).style = header_style

            # set index style
            if index:
                if j < index_nlevels:
                    if index_style == "default":
                        ws.cell(row_num,col_num).font = Font(bold=True)
                    elif index_style is None:
                        ws.cell(row_num,col_num).style = default_style
                    else:
                        ws.cell(row_num,col_num).style = index_style

            
            # set cell style
            if i >= header_nlevels and j >= index_nlevels:
                # apply style by column
                if style_axis == 1:
                    current_style = next(styles_cycle)
                if styles:
                    ws.cell(row_num,col_num).style = current_style

    # table range:
    min_col = col_anchor
    min_row = row_anchor
    max_col = min_col + j
    max_row = min_row + i
    n_header = header_nlevels + int(index) if index else header_nlevels
    table_range = TableRange(ws,min_col=min_col,min_row=min_row,max_col=max_col,
        max_row=max_row,n_index=index_nlevels,n_header=n_header)
    return(table_range)

# alias for functions
write_dataframe_simple = df_to_sheet_simple
write_dataframe_minimal = df_to_sheet_minimal