from openpyxlplus import cell_range
from numpy import ndarray,array,newaxis
from openpyxl.styles import Alignment

def write_value(data,ws,cell=None,keep_style=True):
    """
    Write value to a cell.

    Parameters:
    data: data to write to cell.
    ws: openpyxl.worksheet.worksheet.Worksheet object
    cell: openpyxl.cell.cell.Cell object. Default to A1 cell of ws.
    keep_style: Whether to preserve original cell style

    Return:
    Returns SheetCellRange object, which allows modification to where the data 
        is written to.
    """
    if cell is None:
        cell = ws["A1"]
    rg = cell_range.SheetCellRange(ws,range_string=cell.coordinate)\
        .write(data,keep_style=keep_style)
    return(rg)
    

def write_list(data,ws,cell=None,direction="down",keep_style=True):
    """
    Write data to row or column

    Parameters:
    data: list,tuple or other iterable objects. Does NOT support nested list.
    ws: openpyxl.worksheet.worksheet.Worksheet object
    cell: openpyxl.cell.cell.Cell object. Cell to start writing. Default to 
        A1 cell of ws.
    direction: one of "up","down","left" or "right". From the starting cell, 
        direction to write data to.
    keep_style: Whether to preserve original cell styles.

    Return:
    Returns SheetCellRange object, which allows modification to where the data 
        is written to.
    """
    if cell is None:
        cell = ws["A1"]
    length = len(data)
    row, col = cell.row, cell.column
    if direction == "down":
        min_col = col
        min_row = row
        max_col = col
        max_row = row+length-1
        data = array(data)[newaxis].transpose()
    elif direction == "up":
        min_col = col
        min_row = row-length+1
        if min_row < 1:
            raise ValueError("Min row is fewer than 1. Increase starting row.")
        max_col = col
        max_row = row
        data = data[::-1]
        data = array(data)[newaxis].transpose()
    elif direction == "right":
        min_col = col
        min_row = row
        max_col = col + length - 1
        max_row = row
    elif direction == "left":
        min_col = col - length + 1
        if min_col < 1:
            raise ValueError(
                "Min column is fewer than 1. Increase starting column."
            )
        min_row = row
        max_col = col
        max_row = row
        data = data[::-1]
    else:
        raise ValueError(f"{direction} is not one of up,down,left or right.")
    rg = cell_range.SheetCellRange(ws,min_col=min_col,min_row=min_row,
        max_col=max_col,max_row=max_row).write(data,keep_style=keep_style)
    return(rg)

def write_array(data,ws,cell=None,keep_style=True):
    """
    Write data to range starting at provided cell.

    Parameters:
    data: List of list or numpy array. If data is list of list, length of all 
        nested lists must be equal.
    ws: openpyxl.worksheet.worksheet.Worksheet object
    cell: openpyxl.cell.cell.Cell object. Cell to start writing, this cell is 
        the top left cell in array. Default to first A1 of ws.
    keep_style: Whether to preserve original cell styles.

    Return:
    Returns SheetCellRange object, which allows modification to where the data 
        is written to.
    """
    if cell is None:
        cell = ws["A1"]
    width = len(data[0])
    height = len(data)
    if type(data) != ndarray:
        for i,x in enumerate(data):
            if len(x) != width:
                raise ValueError(
                    f"Length of nested {str(type(x))} mismatch: "
                    f"length of {x} ({len(x)}) is not equal to {width}."
                )

    min_col = cell.column
    min_row = cell.row
    max_col = cell.column + width - 1
    max_row = cell.row + height - 1

    rg = cell_range.SheetCellRange(ws,min_col=min_col,min_row=min_row,
        max_col=max_col,max_row=max_row).write(data,keep_style=keep_style)
    return(rg)


def write_value_merged(
    data,
    ws,
    cell=None,
    right = 0,
    down = 0,
    left = 0,
    up = 0,
    center=True,
    keep_style=True
):
    """
    Write value to given cell and merge cells with shape defined by shape
    parameter

    Parameters:
    data: scalar value to write to worksheet
    ws: openpyxl.worksheet.worksheet.Worksheet object
    cell: openpyxl.cell.cell.Cell object. Cell to start writing, this cell is 
        the top left cell in array. Default to first A1 of ws.
    right,down,left,up: How many cells to expand in given direction.
    center: Whether to center value in merged cell range
    keep_style: Whether to preserve original cell styles.
    """
    if cell is None:
        cell = ws["A1"]
    rg = cell_range.SheetCellRange(ws,range_string=cell.coordinate)\
        .write(data,keep_style=keep_style)

    rg.expand(right=right,down=down,left=left,up=up)
    rg.cells.set_value(data)
    if center:
        rg.cells.modify_style(
            "alignment",
            Alignment(horizontal='center',vertical="center")
        )
    rg.merge_cells()
    return(rg)

def write_dataframe(
        data,
        ws,
        cell=None,
        index=True,
        header=True,
        keep_style=True,
        merge_header=False,
        merge_index=False
    ):
    """
    Write pandas data frame to range starting at provided cell.

    Parameters:
    data: pandas dataframe.
    ws: openpyxl.worksheet.worksheet.Worksheet object
    cell: openpyxl.cell.cell.Cell object. Cell to start writing, this cell is 
        the top left cell in array. Default to first A1 of ws.
    index: Whether to write index
    header: Whether to write header
    keep_style: Whether to preserve original cell styles.

    Return:
    Returns TableRange object, which allows modification to where the data is 
        written to.
    """
    # separate tables to write into 3 parts: header, index and body
    # write these 3 parts separtely
    if cell is None:
        cell = ws["A1"]
    if index:
        index_levels = len(data.index.names)
        if index_levels == 1:
            index_to_write = data.index.to_numpy().reshape((-1,1))
        else:
            index_to_write = array([list(x) for x in data.index.to_numpy()])
    else:
        index_to_write = None
        index_levels = 0
        
    if header:
        header_levels = len(data.columns.names)
        if header_levels == 1:
            header_to_write = data.columns.to_numpy().reshape((1,-1))
        else:
            header_to_write = array([list(x) for x in data.columns.to_numpy().tolist()]).transpose()
    else:
        header_to_write = None
        header_levels = 0
        
    body_to_write = data.to_numpy()
    
    anchor_row_master, anchor_col_master = cell.row, cell.col_idx
    anchor_row_index, anchor_col_index = cell.row, cell.col_idx
    anchor_row_header, anchor_col_header =cell.row, cell.col_idx
    anchor_row_body, anchor_col_body = cell.row, cell.col_idx
    
    if header:
        anchor_row_index += header_levels
        anchor_row_body +=  header_levels
        
    if index:
        anchor_col_header += index_levels
        anchor_col_body += index_levels
    
    # write header if provided
    if header:
        rg = write_array(
            data = header_to_write,
            ws = ws,
            cell = ws.cell(anchor_row_header,anchor_col_header),
            keep_style = keep_style
        )

        if merge_header:
            rg.merge_consecutive_cells(on="row")
    
    # write index if provided
    if index:
        rg = write_array(
            data = index_to_write,
            ws = ws,
            cell = ws.cell(anchor_row_index,anchor_col_index),
            keep_style = keep_style
        )

        if merge_index:
            rg.merge_consecutive_cells(on="column")

    # write body
    write_array(
        data = body_to_write,
        ws = ws,
        cell = ws.cell(anchor_row_body,anchor_col_body),
        keep_style = keep_style
    )
    
    # create TableRange object
    table_range = cell_range.TableRange(
        ws,
        min_col=anchor_col_master,
        min_row=anchor_row_master,
        max_col=anchor_col_master + index_levels + body_to_write.shape[1] - 1, 
        max_row=anchor_row_master + header_levels + body_to_write.shape[0] - 1,
        n_index=index_levels,
        n_header=header_levels
    )
    return(table_range)


# def _merge_consecutive_cells(rg,on="row",center=True):
#     """
#     Merge consecutive cells by rows or by columns

#     Parameters:
#     rg: openpyxlplus.cell_range.SheetCellRange
#     on: "row" or "column". "row" to merge horizontally (same columns are merged);
#         "column" to merge vertically (different columns are merged)
#     center: True to center merged cells.
#     """
#     if on == "row":
#         temp = rg.cells
#     elif on == "column":
#         temp = rg.cells.transpose()
#     else:
#         raise Exception(f"{on} not supported.")
#     groups = []
#     for row in temp:
#         g = []
#         for cell in row:
#             if len(g) == 0:
#                 g.append(cell)
#             else:
#                 if cell.value == g[-1].value:
#                     g.append(cell)
#                 else:
#                     groups.append(g)
#                     g = [cell]
#         groups.append(g)

#     for g in groups:
#         if len(g) > 1:
#             rg_temp = cell_range.Cells(g).to_range()
#             if center:
#                 rg_temp.cells.modify_style(
#                     "alignment",
#                     Alignment(horizontal='center',vertical="center")
#                 )
#             rg_temp.merge_cells()
#     return(rg)